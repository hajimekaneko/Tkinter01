import json
import logging
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

def extract_xlsx_to_json(file_path: str) -> dict:
    """
    指定された XLSX ファイルの内容を JSON 形式の辞書として抜き出す関数。
    
    ・各シートの 1 行目をヘッダーとして使用し、2 行目以降をデータ行として扱う。
    ・ヘッダーに「グループ」が存在すれば、その列の値をグループキーとして利用し、
      グループ列が空欄の場合は直前のグループ（current_group）の値を割り当てる。
    ・ただし、行のうち「グループ」以外のすべての値が null の場合は、
      グループ更新用の行とみなし、データとしては出力しない。
    
    結果は { シート名: { グループ名: [rowの辞書, ...], ... } } の形式となる。
    """
    try:
        workbook = load_workbook(file_path, data_only=True)
        result = {}
        for sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                result[sheet_name] = {}
                continue
            headers = list(rows[0])
            try:
                group_index = headers.index("グループ")
            except ValueError:
                group_index = 0

            data_by_group = {}
            current_group = None

            for row in rows[1:]:
                if all(cell is None for cell in row):
                    continue
                row_dict = {header: value for header, value in zip(headers, row)}
                group_value = row[group_index]
                # 更新: グループ列に値があれば current_group を更新
                if group_value is not None and str(group_value).strip() != "":
                    current_group = group_value
                if current_group is None:
                    current_group = "UNDEFINED"
                
                # 判定: 「グループ」以外の全列が None なら、これはグループ変更用行とみなす
                is_group_header = False
                if group_value is not None and str(group_value).strip() != "":
                    # グループ列以外の値を取得
                    other_values = [row_dict[header] for i, header in enumerate(headers) if i != group_index]
                    if all(val is None for val in other_values):
                        is_group_header = True
                
                # 常に「グループ」フィールドは current_group に設定
                row_dict["グループ"] = current_group
                
                # グループ変更用行の場合は、出力対象から除外する
                if is_group_header:
                    continue

                if current_group not in data_by_group:
                    data_by_group[current_group] = []
                data_by_group[current_group].append(row_dict)
            result[sheet_name] = data_by_group
        return result
    except Exception as e:
        logger.error("XLSX ファイルの読み込みエラー: %s", e)
        raise e

def extract_xlsx_to_json_str(file_path: str) -> str:
    data = extract_xlsx_to_json(file_path)
    return json.dumps(data, ensure_ascii=False, indent=4)
